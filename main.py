import math
import openpyxl
from openpyxl import Workbook

def baca_data(nama_file):
    """Membaca data dari file Excel dan mengembalikan dictionary data restoran"""
    wb = openpyxl.load_workbook(nama_file)
    sheet = wb.active
    
    data_restoran = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        id_pelanggan = int(row[0])
        servis = int(row[1])
        harga = float(row[2])
        data_restoran[id_pelanggan] = {'servis': servis, 'harga': harga}
    
    return data_restoran

def fuzzifikasi_servis(servis):
    """Melakukan fuzzifikasi untuk atribut kualitas servis"""
    # Fungsi keanggotaan untuk servis: Buruk, Cukup, Baik, Sangat Baik
    buruk = max(0, min(1, (30 - servis) / 30)) if servis <= 30 else 0
    cukup = max(0, min((servis - 20) / 20, (60 - servis) / 20)) if 20 <= servis <= 60 else 0
    baik = max(0, min((servis - 50) / 20, (80 - servis) / 20)) if 50 <= servis <= 80 else 0
    sangat_baik = max(0, min(1, (servis - 70) / 30)) if servis >= 70 else 0
    
    return {'Buruk': buruk, 'Cukup': cukup, 'Baik': baik, 'Sangat Baik': sangat_baik}

def fuzzifikasi_harga(harga):
    """Melakukan fuzzifikasi untuk atribut harga"""
    # Fungsi keanggotaan untuk harga: Murah, Sedang, Mahal
    murah = max(0, min(1, (35000 - harga) / 15000)) if harga <= 35000 else 0
    sedang = max(0, min((harga - 25000) / 15000, (50000 - harga) / 15000)) if 25000 <= harga <= 50000 else 0
    mahal = max(0, min(1, (harga - 40000) / 15000)) if harga >= 40000 else 0
    
    return {'Murah': murah, 'Sedang': sedang, 'Mahal': mahal}

def inferensi(fuzzy_servis, fuzzy_harga):
    """Melakukan inferensi berdasarkan aturan fuzzy"""
    # Aturan fuzzy:
    # 1. IF Servis Buruk OR Harga Mahal THEN Tidak Layak
    # 2. IF Servis Cukup AND Harga Sedang THEN Cukup Layak
    # 3. IF Servis Baik AND Harga Murah THEN Layak
    # 4. IF Servis Sangat Baik AND Harga Murah THEN Sangat Layak
    
    # Rule 1: Buruk OR Mahal -> Tidak Layak
    tidak_layak = max(fuzzy_servis['Buruk'], fuzzy_harga['Mahal'])
    
    # Rule 2: Cukup AND Sedang -> Cukup Layak
    cukup_layak = min(fuzzy_servis['Cukup'], fuzzy_harga['Sedang'])
    
    # Rule 3: Baik AND Murah -> Layak
    layak = min(fuzzy_servis['Baik'], fuzzy_harga['Murah'])
    
    # Rule 4: Sangat Baik AND Murah -> Sangat Layak
    sangat_layak = min(fuzzy_servis['Sangat Baik'], fuzzy_harga['Murah'])
    
    return {
        'Tidak Layak': tidak_layak,
        'Cukup Layak': cukup_layak,
        'Layak': layak,
        'Sangat Layak': sangat_layak
    }

def defuzzifikasi(fuzzy_output):
    """Melakukan defuzzifikasi menggunakan metode centroid"""
    # Fungsi keanggotaan output (kelayakan)
    # Tidak Layak: 0-30
    # Cukup Layak: 20-50
    # Layak: 40-70
    # Sangat Layak: 60-100
    
    # Parameter untuk defuzzifikasi
    step = 1
    total_area = 0
    weighted_sum = 0
    
    for x in range(0, 101, step):
        # Hitung nilai keanggotaan untuk setiap x
        tidak_layak = max(0, min(1, (30 - x) / 30)) if x <= 30 else 0
        cukup_layak = max(0, min((x - 20) / 15, (50 - x) / 15)) if 20 <= x <= 50 else 0
        layak = max(0, min((x - 40) / 15, (70 - x) / 15)) if 40 <= x <= 70 else 0
        sangat_layak = max(0, min((x - 60) / 40, (100 - x) / 40)) if 60 <= x <= 100 else 0
        
        # Gabungkan dengan fuzzy output menggunakan operasi MAX
        membership = max(
            min(tidak_layak, fuzzy_output['Tidak Layak']),
            min(cukup_layak, fuzzy_output['Cukup Layak']),
            min(layak, fuzzy_output['Layak']),
            min(sangat_layak, fuzzy_output['Sangat Layak'])
        )
        
        total_area += membership
        weighted_sum += x * membership
    
    if total_area == 0:
        return 0  # Hindari pembagian dengan nol
    
    return weighted_sum / total_area

def proses_fuzzy(data_restoran):
    """Memproses semua data restoran dengan fuzzy logic"""
    hasil = []
    
    for id_restoran, atribut in data_restoran.items():
        servis = atribut['servis']
        harga = atribut['harga']
        
        # Fuzzifikasi
        fuzzy_servis = fuzzifikasi_servis(servis)
        fuzzy_harga = fuzzifikasi_harga(harga)
        
        # Inferensi
        fuzzy_output = inferensi(fuzzy_servis, fuzzy_harga)
        
        # Defuzzifikasi
        skor = defuzzifikasi(fuzzy_output)
        
        hasil.append({
            'id': id_restoran,
            'servis': servis,
            'harga': harga,
            'skor': skor
        })
    
    return hasil

def simpan_hasil(nama_file, hasil):
    """Menyimpan hasil ke file Excel"""
    wb = Workbook()
    ws = wb.active
    
    # Header
    ws.append(['ID Restoran', 'Kualitas Servis', 'Harga', 'Skor Kelayakan'])
    
    # Data
    for item in hasil:
        ws.append([item['id'], item['servis'], item['harga'], item['skor']])
    
    wb.save(nama_file)

def main():
    # Baca data dari file
    data_restoran = baca_data('data/restoran.xlsx')
    
    # Proses fuzzy logic
    hasil_fuzzy = proses_fuzzy(data_restoran)
    
    # Urutkan berdasarkan skor tertinggi
    hasil_terurut = sorted(hasil_fuzzy, key=lambda x: x['skor'], reverse=True)
    
    # Ambil 10 terbaik
    top_10 = hasil_terurut[:10]
    
    # Simpan hasil ke file
    simpan_hasil('data/peringkat.xlsx', top_10)
    
    # Tampilkan hasil
    print("10 Restoran Terbaik:")
    print("{:<10} {:<15} {:<15} {:<15}".format("ID", "Kualitas Servis", "Harga", "Skor"))
    for restoran in top_10:
        print("{:<10} {:<15} {:<15.2f} {:<15.2f}".format(
            restoran['id'], restoran['servis'], restoran['harga'], restoran['skor']))

if __name__ == "__main__":
    main()